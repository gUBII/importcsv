"""
Assist Service Type Variant Truth Extractor

Extracts variant-level data from Assist appointments form:
- All variants per Service Type (not just defaults)
- Complete Rate + Code pairs per variant
- Checkpointing for resumable extraction
- Conflict detection across probe clients

Output: ~/LineItemRates/ServiceTypeTruth/variants/{latest,snapshots}/
"""

import csv
import json
import re
import shutil
import time
import traceback
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple, Union

from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

from importcsv import (
    BASE_URL,
    build_chrome_driver,
    ensure_credentials,
    login,
    log_message,
)
import line_item_paths
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# =============================================================================
# TYPES & CALLBACKS
# =============================================================================

ProgressCallback = Optional[Callable[[str], None]]
EventCallback = Optional[Callable[[Dict[str, str]], None]]
RowCallback = Optional[Callable[[Dict[str, str]], None]]

EXCEL_ILLEGAL_CHAR_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

# =============================================================================
# OUTPUT SCHEMA
# =============================================================================

VARIANT_COLUMNS = [
    "Parent Service Type ID",
    "Parent Service Type Label",
    "Parent Service Type",
    "Service Type ID",
    "Service Variant Label",
    "Rate",
    "Rate (Raw)",
    "Code",
    "Code (Raw)",
    "Item Number",
    "Unit",
    "Status",
    "Error Reason",
    "Conflict",
    "Conflict Detail",
    "Probe Client ID",
    "Source URL",
    "Captured At (UTC)",
]

# =============================================================================
# DIAGNOSTICS FIELDS
# =============================================================================

EVENT_FIELDS = [
    "run_id",
    "ts_utc",
    "level",
    "step",
    "code",
    "message",
    "client_id",
    "url",
    "selector",
    "option_count",
    "context_json",
]

CHECKER_FIELDS = [
    "run_id",
    "ts_utc",
    "step",
    "status",
    "code",
    "message",
    "client_id",
    "url",
    "selector",
    "option_count",
    "artifact_screenshot",
    "artifact_html",
]

# Checker codes
CHECKER_CLIENT = "CHK_CLIENT_PAGE_OPEN"
CHECKER_APPOINTMENT = "CHK_APPOINTMENT_ENTRY_REACHED"
CHECKER_SERVICE_TYPE_SELECT = "CHK_SERVICE_TYPE_SELECT"
CHECKER_VARIANT_TABLE_EXTRACT = "CHK_VARIANT_TABLE_EXTRACT"
CHECKER_VARIANT_TABLE_MISSING = "CHK_VARIANT_TABLE_MISSING"
CHECKER_VARIANT_TABLE_EMPTY = "CHK_VARIANT_TABLE_EMPTY"
CHECKER_VARIANTS_EDITOR_ROUTE = "CHK_VARIANTS_EDITOR_ROUTE"
CHECKER_VARIANTS_EDITOR_CAPABILITY = "CHK_VARIANTS_EDITOR_CAPABILITY"

ASSIST_DIRECT_URL = "https://assist.turnpoint.co/appointments/new"
SERVICE_TYPE_INPUT_XPATH = (
    "//div[@data-cy='service_type_id-reload']/preceding::input[@role='combobox'][1]"
)
SERVICE_TYPE_RELOAD_CSS = "div[data-cy='service_type_id-reload']"
SERVICE_TYPE_LISTBOX_CSS = "div[role='listbox']"
SERVICE_TYPE_OPTION_XPATH = "//div[@role='listbox']//div[@role='option']"
OUTER_APPOINTMENT_IFRAME_CSS = "iframe[src*='appointment-edit.asp']"
INNER_ASSIST_IFRAME_CSS = (
    "iframe[src*='assist.turnpoint.co/appointments/new'][src*='has_parent=true']"
)
VARIANT_PREFIXES = ["day", "eve", "night", "saturday", "sunday", "ph"]
VARIANT_LABEL_FALLBACK = {
    "day": "Weekday Daytime/Individual Code",
    "eve": "Weekday Evening",
    "night": "Weekday Night",
    "saturday": "Saturday",
    "sunday": "Sunday",
    "ph": "Public Holiday",
}

MODAL_DISMISS_BUTTON_XPATH = (
    "//button[normalize-space()='Close' or normalize-space()='Dismiss' or normalize-space()='OK']"
    " | //button[contains(normalize-space(.), 'close') or contains(normalize-space(.), 'Dismiss')]"
    " | //div[@role='dialog']//button"
)

# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================


def _utc_now() -> str:
    """Return current UTC timestamp in ISO format."""
    return datetime.now(timezone.utc).isoformat() + "Z"


def _emit(message: str, callback: ProgressCallback = None):
    """Emit progress message."""
    if callback:
        callback(message)


def _normalize_text(value) -> str:
    """Normalize text: strip, collapse whitespace."""
    if not value:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def _normalize_token(value: str) -> str:
    """Normalize token: strip, lowercase, alphanumeric + underscore."""
    text = _normalize_text(value).lower()
    return re.sub(r"[^a-z0-9_]", "", text)


def _safe_json(value) -> str:
    """Safely serialize value to JSON."""
    try:
        return json.dumps(value, default=str)
    except Exception:
        return str(value)


def _coerce_rate_text(value: str) -> str:
    """Extract and normalize numeric rate."""
    if not value:
        return ""
    text = _normalize_text(value)
    matches = re.findall(r"\d[\d,]*\.?\d*", text)
    if not matches:
        return text
    return matches[0].replace(",", "")


def _normalize_code_text(value: str) -> str:
    """Normalize item code while preserving underscore separators."""
    if not value:
        return ""
    cleaned = re.sub(r"[^A-Za-z0-9_]", "", _normalize_text(value))
    return cleaned.upper()


def _split_rate_and_unit(value: str, unit_hint: str = "") -> Tuple[str, str]:
    """Split a rate cell into normalized numeric rate and raw unit text."""
    raw = _normalize_text(value)
    rate = _coerce_rate_text(raw)
    unit = _normalize_text(unit_hint)
    if not unit and "/" in raw:
        unit = _normalize_text(raw[raw.index("/") :])
    return rate, unit


def _is_interactable(element) -> bool:
    """Best-effort interactable check that also works with test doubles."""
    if element is None:
        return False
    try:
        displayed = element.is_displayed()
    except Exception:
        displayed = True
    try:
        enabled = element.is_enabled()
    except Exception:
        enabled = True
    return bool(displayed and enabled)


def _sanitize_excel_text(value) -> Tuple[str, int]:
    """Remove XLSX-illegal characters."""
    if not value:
        return "", 0
    text = str(value)
    cleaned = EXCEL_ILLEGAL_CHAR_RE.sub("", text)
    return cleaned, len(text) - len(cleaned)


# =============================================================================
# DIAGNOSTICS RECORDER
# =============================================================================


class DiagnosticsRecorder:
    """Record diagnostics events and artifacts."""

    def __init__(
        self,
        run_id: str,
        output_dir: Path,
        on_event: EventCallback = None,
    ):
        self.run_id = run_id
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.events: List[Dict[str, str]] = []
        self.checkers: List[Dict[str, str]] = []
        self._event_callback = on_event

    def event(
        self,
        level: str,
        step: str,
        code: str,
        message: str,
        client_id: Optional[str] = None,
        url: Optional[str] = None,
        selector: Optional[str] = None,
        option_count: Optional[int] = None,
        context: Optional[Dict[str, Any]] = None,
    ):
        """Record event."""
        event_dict = {
            "run_id": self.run_id,
            "ts_utc": _utc_now(),
            "level": level,
            "step": step,
            "code": code,
            "message": message,
            "client_id": client_id or "",
            "url": url or "",
            "selector": selector or "",
            "option_count": str(option_count or ""),
            "context_json": _safe_json(context) if context else "",
        }
        self.events.append(event_dict)
        if self._event_callback:
            try:
                self._event_callback(dict(event_dict))
            except Exception:
                pass

    def checker(
        self,
        step: str,
        status: str,
        code: str,
        message: str,
        client_id: Optional[str] = None,
        url: Optional[str] = None,
        selector: Optional[str] = None,
        option_count: Optional[int] = None,
        screenshot: Optional[Path] = None,
        html: Optional[Path] = None,
    ):
        """Record checker result."""
        checker_dict = {
            "run_id": self.run_id,
            "ts_utc": _utc_now(),
            "step": step,
            "status": status,
            "code": code,
            "message": message,
            "client_id": client_id or "",
            "url": url or "",
            "selector": selector or "",
            "option_count": str(option_count or ""),
            "artifact_screenshot": str(screenshot.name) if screenshot else "",
            "artifact_html": str(html.name) if html else "",
        }
        self.checkers.append(checker_dict)

    def save(self):
        """Save events and checkers to disk."""
        # Save events as JSONL
        events_path = self.output_dir / "events.jsonl"
        with open(events_path, "w", encoding="utf-8") as f:
            for event in self.events:
                f.write(json.dumps(event) + "\n")

        # Save checkers as CSV
        checkers_path = self.output_dir / "checkers.csv"
        with open(checkers_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=CHECKER_FIELDS)
            writer.writeheader()
            if self.checkers:
                writer.writerows(self.checkers)

    def capture_screenshot(self, driver, name: str) -> Optional[Path]:
        """Capture screenshot and save."""
        path = self.output_dir / f"{name}.png"
        try:
            driver.save_screenshot(str(path))
            return path
        except Exception:
            return None

    def capture_html(self, driver, name: str) -> Optional[Path]:
        """Capture page HTML and save."""
        path = self.output_dir / f"{name}.html"
        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            return path
        except Exception:
            return None


# =============================================================================
# FILE I/O
# =============================================================================


def _write_csv(rows: List[Dict[str, str]], fieldnames: List[str], path: Path):
    """Write rows to CSV file."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            # Sanitize Excel-illegal chars
            cleaned_row = {}
            for key in fieldnames:
                value = row.get(key, "")
                if value is not None:
                    cleaned, _ = _sanitize_excel_text(value)
                    cleaned_row[key] = cleaned
                else:
                    cleaned_row[key] = ""
            writer.writerow(cleaned_row)


def _write_xlsx(rows: List[Dict[str, str]], fieldnames: List[str], path: Path) -> Path:
    """Write rows to XLSX file."""
    import openpyxl

    path.parent.mkdir(parents=True, exist_ok=True)

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write header
    for col_idx, field in enumerate(fieldnames, 1):
        ws.cell(1, col_idx).value = field

    # Write rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, field in enumerate(fieldnames, 1):
            value = row.get(field, "")
            if value is not None:
                cleaned, _ = _sanitize_excel_text(value)
                ws.cell(row_idx, col_idx).value = cleaned

    wb.save(str(path))
    return path


def _apply_xlsx_merged_cells(xlsx_path: Path):
    """Apply merged cells and formatting to XLSX file."""
    try:
        wb = load_workbook(xlsx_path)
        ws = wb.active

        # Merge contiguous parent groups in columns A and B.
        group_start = None
        current_parent = None
        for row_idx in range(2, ws.max_row + 2):
            parent_id = ws.cell(row_idx, 1).value if row_idx <= ws.max_row else None
            if parent_id != current_parent:
                if current_parent and group_start and row_idx - group_start > 1:
                    end_row = row_idx - 1
                    try:
                        ws.merge_cells(f"A{group_start}:A{end_row}")
                        ws.merge_cells(f"B{group_start}:B{end_row}")
                    except Exception:
                        pass
                    ws.cell(group_start, 1).alignment = Alignment(vertical="center")
                    ws.cell(group_start, 2).alignment = Alignment(vertical="center")
                current_parent = parent_id
                group_start = row_idx if parent_id else None

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(str(xlsx_path))
    except Exception as e:
        log_message(f"Warning: Could not apply XLSX merges: {e}")


# =============================================================================
# CHECKPOINT MANAGEMENT
# =============================================================================


def _load_checkpoint(run_id: str) -> Optional[Dict[str, Any]]:
    """Load checkpoint JSON from disk."""
    paths = line_item_paths.get_variant_paths(run_id)
    checkpoint_path = paths["checkpoint_json"]

    if not checkpoint_path.exists():
        return None

    try:
        with open(checkpoint_path, "r") as f:
            return json.load(f)
    except Exception:
        return None


def _update_checkpoint(run_id: str, checkpoint: Dict[str, Any]):
    """Save checkpoint JSON to disk."""
    paths = line_item_paths.get_variant_paths(run_id)
    checkpoint_path = paths["checkpoint_json"]
    checkpoint_path.parent.mkdir(parents=True, exist_ok=True)

    with open(checkpoint_path, "w") as f:
        json.dump(checkpoint, f, indent=2)


def _append_to_checkpoint_csv(run_id: str, rows: List[Dict[str, str]]):
    """Append rows to checkpoint CSV buffer."""
    if not rows:
        return

    paths = line_item_paths.get_variant_paths(run_id)
    csv_path = paths["checkpoint_append_csv"]
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    # Check if file exists
    file_exists = csv_path.exists()

    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=VARIANT_COLUMNS)
        if not file_exists:
            writer.writeheader()
        for row in rows:
            cleaned_row = {}
            for key in VARIANT_COLUMNS:
                value = row.get(key, "")
                if value is not None:
                    cleaned, _ = _sanitize_excel_text(value)
                    cleaned_row[key] = cleaned
                else:
                    cleaned_row[key] = ""
            writer.writerow(cleaned_row)


# =============================================================================
# ASSIST NAVIGATION & EXTRACTION (Reused patterns)
# =============================================================================


def _switch_to_default_content(driver) -> None:
    """Switch to top-level document when Selenium context supports it."""
    try:
        if hasattr(driver, "switch_to"):
            driver.switch_to.default_content()
    except Exception:
        pass


def _switch_to_frame(driver, frame_elem) -> bool:
    """Switch into a frame if possible."""
    try:
        if hasattr(driver, "switch_to"):
            driver.switch_to.frame(frame_elem)
            return True
    except Exception:
        return False
    return False


def _record_locator_success(
    recorder: DiagnosticsRecorder,
    *,
    step: str,
    code: str,
    strategy: str,
    selector: str,
    frame_hint: str,
) -> None:
    recorder.event(
        "INFO",
        step,
        code,
        f"Locator strategy '{strategy}' matched in {frame_hint}",
        selector=selector,
        context={"strategy": strategy, "frame": frame_hint},
    )


def _capture_assist_failure_artifacts(
    driver,
    recorder: DiagnosticsRecorder,
    *,
    prefix: str,
    reason: str,
) -> Tuple[Optional[Path], Optional[Path]]:
    """Capture screenshot + HTML + console logs for failure forensics."""
    safe_prefix = _normalize_token(prefix) or "assist"
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    screenshot = recorder.capture_screenshot(driver, f"{safe_prefix}_{ts}")
    html = recorder.capture_html(driver, f"{safe_prefix}_{ts}")

    current_url = _normalize_text(getattr(driver, "current_url", ""))
    title = _normalize_text(getattr(driver, "title", ""))
    body_len = 0
    body_snippet_len = 0
    try:
        body = _normalize_text(driver.find_element(By.TAG_NAME, "body").text)
        body_len = len(body)
        body_snippet_len = len(body[:500])
    except Exception:
        pass

    console_path = recorder.output_dir / f"{safe_prefix}_{ts}_console.json"
    console_saved = False
    try:
        logs = driver.get_log("browser")
        with open(console_path, "w", encoding="utf-8") as fh:
            json.dump(logs, fh, indent=2, default=str)
        console_saved = True
    except Exception:
        console_saved = False

    recorder.event(
        "ERROR",
        "assist_artifacts",
        "ASSIST_FAILURE_ARTIFACTS",
        reason,
        url=current_url,
        context={
            "current_url": current_url,
            "title": title,
            "body_len": body_len,
            "body_snippet_len": body_snippet_len,
            "html_artifact": html.name if html else "",
            "screenshot_artifact": screenshot.name if screenshot else "",
            "console_artifact": console_path.name if console_saved else "",
        },
    )
    return screenshot, html


def _dismiss_blocking_modal_if_present(driver, recorder: DiagnosticsRecorder) -> bool:
    """Dismiss an obvious blocking modal when present."""
    dismissed = False
    for strategy_name, by, selector in [
        ("modal-close-buttons", By.XPATH, MODAL_DISMISS_BUTTON_XPATH),
        ("modal-role-dialog", By.CSS_SELECTOR, "[role='dialog'] button"),
        ("bootstrap-modal", By.CSS_SELECTOR, ".modal button"),
    ]:
        try:
            buttons = driver.find_elements(by, selector)
        except Exception:
            buttons = []
        for button in buttons:
            if not _is_interactable(button):
                continue
            try:
                button.click()
                dismissed = True
                recorder.event(
                    "INFO",
                    "dismiss_modal",
                    "MODAL_DISMISSED",
                    f"Dismissed blocking modal using {strategy_name}",
                    selector=selector,
                    context={"strategy": strategy_name},
                )
                time.sleep(0.2)
                return True
            except Exception:
                continue
    return dismissed


def _find_service_type_input(driver, require_interactable: bool = True):
    """Locate the Service Type combobox using the Atlas-confirmed primary XPath."""
    try:
        candidates = driver.find_elements(By.XPATH, SERVICE_TYPE_INPUT_XPATH)
    except Exception:
        candidates = []
    for elem in candidates:
        if not require_interactable or _is_interactable(elem):
            return elem
    return None


def _wait_for_listbox_open(driver, timeout: float = 4.0) -> bool:
    """Wait for listbox open state."""
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            listboxes = driver.find_elements(By.CSS_SELECTOR, SERVICE_TYPE_LISTBOX_CSS)
        except Exception:
            listboxes = []
        if listboxes:
            return True
        time.sleep(0.15)
    return False


def _wait_for_combobox_closed(input_el, timeout: float = 5.0) -> bool:
    """Wait for aria-expanded=false after selection."""
    deadline = time.time() + timeout
    while time.time() < deadline:
        expanded = _normalize_text(input_el.get_attribute("aria-expanded")).lower()
        if expanded in ("false", ""):
            return True
        time.sleep(0.12)
    return False


def _clear_and_type(input_el, text: str):
    """Clear combobox and type query text."""
    try:
        input_el.click()
    except Exception:
        pass
    try:
        input_el.clear()
    except Exception:
        pass
    try:
        input_el.send_keys(Keys.CONTROL, "a")
        input_el.send_keys(Keys.BACKSPACE)
    except Exception:
        pass
    if text:
        input_el.send_keys(text)


def _click_matching_service_option(driver, expected_label: str, allow_contains: bool = False) -> bool:
    """Click a matching option in div[role='listbox'] list."""
    try:
        options = driver.find_elements(By.XPATH, SERVICE_TYPE_OPTION_XPATH)
    except Exception:
        options = []
    if not options:
        return False

    target = _normalize_text(expected_label).lower()
    for option in options:
        text = _normalize_text(getattr(option, "text", ""))
        if text.lower() == target:
            option.click()
            return True

    if allow_contains and target:
        for option in options:
            text = _normalize_text(getattr(option, "text", ""))
            if target in text.lower():
                option.click()
                return True

    return False


def _select_service_type_option(
    driver,
    query_text: str,
    expected_label: str,
    recorder: DiagnosticsRecorder,
    *,
    allow_contains: bool = False,
) -> bool:
    """Select service type by type+enter, then exact option-click fallback."""
    input_el = _find_service_type_input(driver, require_interactable=True)
    if not input_el:
        return False

    _clear_and_type(input_el, query_text)
    _wait_for_listbox_open(driver, timeout=4.0)
    try:
        input_el.send_keys(Keys.ENTER)
    except Exception:
        pass

    if _wait_for_combobox_closed(input_el, timeout=4.0):
        return True

    # Fallback to explicit option click by visible text.
    _clear_and_type(input_el, query_text)
    if _wait_for_listbox_open(driver, timeout=3.0):
        clicked = _click_matching_service_option(driver, expected_label, allow_contains=allow_contains)
        if clicked and _wait_for_combobox_closed(input_el, timeout=4.0):
            return True

    recorder.event(
        "WARN",
        "select_option",
        "SERVICE_TYPE_SELECT_NOT_CLOSED",
        f"Combobox did not settle for '{expected_label}'",
    )
    return False


def _variant_input_elements(driver, prefix: str) -> Tuple[Optional[Any], Optional[Any]]:
    """Return (rate_el, code_el) for one prefix."""
    try:
        rate_elems = driver.find_elements(By.CSS_SELECTOR, f"input[data-cy='{prefix}_rate-input']")
    except Exception:
        rate_elems = []
    try:
        code_elems = driver.find_elements(By.CSS_SELECTOR, f"input[data-cy='{prefix}_code-input']")
    except Exception:
        code_elems = []
    rate_el = rate_elems[0] if rate_elems else None
    code_el = code_elems[0] if code_elems else None
    return rate_el, code_el


def _read_variant_fingerprint(driver) -> Tuple[Tuple[str, str, str], ...]:
    """Fingerprint current 6-pair values to detect stale selections."""
    fp: List[Tuple[str, str, str]] = []
    for prefix in VARIANT_PREFIXES:
        rate_el, code_el = _variant_input_elements(driver, prefix)
        rate_val = _normalize_text(rate_el.get_attribute("value")) if rate_el else ""
        code_val = _normalize_text(code_el.get_attribute("value")) if code_el else ""
        fp.append((prefix, rate_val, code_val))
    return tuple(fp)


def _wait_for_fingerprint_change(
    driver,
    before: Tuple[Tuple[str, str, str], ...],
    timeout: float = 6.0,
) -> bool:
    """Wait until variant fingerprint changes."""
    deadline = time.time() + timeout
    while time.time() < deadline:
        after = _read_variant_fingerprint(driver)
        if after != before:
            time.sleep(0.35)
            return True
        time.sleep(0.15)
    return False


def _derive_variant_label(driver, rate_el, prefix: str) -> str:
    """Derive label from nearest row wrapper text before '$', with static fallback."""
    script = """
    const input = arguments[0];
    let node = input;
    for (let i = 0; i < 10 && node; i++) {
      if (node.tagName && node.tagName.toLowerCase() === "div") {
        const txt = (node.innerText || "").replace(/\\s+/g, " ").trim();
        if (txt && txt.includes("$")) {
          const parts = txt.split("$");
          if (parts.length > 0) {
            const label = parts[0].replace(/\\s+/g, " ").trim();
            if (label) return label;
          }
        }
      }
      node = node.parentElement;
    }
    return "";
    """
    try:
        label = _normalize_text(driver.execute_script(script, rate_el))
    except Exception:
        label = ""
    return label or VARIANT_LABEL_FALLBACK.get(prefix, prefix.title())


def _derive_unit_text(driver, rate_el) -> str:
    """Extract '/ hour' unit marker near rate input."""
    script = """
    const input = arguments[0];
    let node = input;
    for (let i = 0; i < 8 && node; i++) {
      const txt = (node.innerText || "").replace(/\\s+/g, " ").trim();
      if (/\\/\\s*hour/i.test(txt)) return "/ hour";
      node = node.parentElement;
    }
    return "";
    """
    try:
        return _normalize_text(driver.execute_script(script, rate_el))
    except Exception:
        return ""


def _extract_variant_table_rows(driver, recorder: DiagnosticsRecorder) -> List[Dict]:
    """
    Extract fixed 6 Service Type variant pairs from data-cy inputs.
    """
    rows: List[Dict[str, str]] = []
    for prefix in VARIANT_PREFIXES:
        rate_el, code_el = _variant_input_elements(driver, prefix)
        if not rate_el or not code_el:
            recorder.event(
                "WARN",
                "extract_variants",
                CHECKER_VARIANT_TABLE_MISSING,
                f"Missing variant input pair for prefix={prefix}",
            )
            return []

        rate_raw = _normalize_text(rate_el.get_attribute("value"))
        code_raw = _normalize_text(code_el.get_attribute("value"))
        rows.append(
            {
                "Service Variant Label": _derive_variant_label(driver, rate_el, prefix),
                "Rate": _coerce_rate_text(rate_raw),
                "Rate (Raw)": rate_raw,
                "Code": _normalize_code_text(code_raw),
                "Code (Raw)": code_raw,
                "Unit": _derive_unit_text(driver, rate_el),
            }
        )

    recorder.checker(
        "extract_variants",
        "PASS",
        CHECKER_VARIANT_TABLE_EXTRACT,
        f"Extracted {len(rows)} rows via fixed variant input pairs",
        option_count=len(rows),
        url=_normalize_text(getattr(driver, "current_url", "")),
    )
    return rows


def _click_add_appointment(driver) -> bool:
    """Click Add Appointment from TP1 client appointments tab."""
    selectors = [
        (By.XPATH, "//a[contains(normalize-space(.), 'Add Appointment')]"),
        (By.XPATH, "//button[contains(normalize-space(.), 'Add Appointment')]"),
        (By.XPATH, "//a[contains(@href, 'appointment-edit.asp')]"),
    ]
    for by, selector in selectors:
        try:
            elems = driver.find_elements(by, selector)
        except Exception:
            elems = []
        for elem in elems:
            if not _is_interactable(elem):
                continue
            try:
                elem.click()
            except Exception:
                try:
                    driver.execute_script("arguments[0].click();", elem)
                except Exception:
                    continue
            return True
    return False


def _switch_to_variants_capable_editor(
    driver,
    client_id: str,
    recorder: DiagnosticsRecorder,
) -> bool:
    """Navigate TP1 Add Appointment flow and switch into nested iframes."""
    if not client_id:
        recorder.checker(
            "open_assist",
            "FAIL",
            CHECKER_VARIANTS_EDITOR_ROUTE,
            "Probe client id is required for variants-capable route.",
            url=_normalize_text(getattr(driver, "current_url", "")),
        )
        return False

    _switch_to_default_content(driver)
    target_url = (
        f"{BASE_URL.rstrip('/')}/client-details.asp?eid={client_id}&BREAKDOWN_SHOW_APPTS=yes&wide1=yes"
    )
    try:
        driver.get(target_url)
        WebDriverWait(driver, 25).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        if not _click_add_appointment(driver):
            raise RuntimeError("Could not click Add Appointment")

        time.sleep(0.6)
        outer = None
        deadline = time.time() + 12
        while time.time() < deadline and not outer:
            try:
                outer_frames = driver.find_elements(By.CSS_SELECTOR, OUTER_APPOINTMENT_IFRAME_CSS)
            except Exception:
                outer_frames = []
            outer = outer_frames[0] if outer_frames else None
            if not outer:
                time.sleep(0.2)
        if not outer:
            raise RuntimeError("Outer appointment-edit iframe not found")

        _switch_to_default_content(driver)
        if not _switch_to_frame(driver, outer):
            raise RuntimeError("Failed to switch to outer iframe")

        inner = None
        deadline = time.time() + 12
        while time.time() < deadline and not inner:
            try:
                inner_frames = driver.find_elements(By.CSS_SELECTOR, INNER_ASSIST_IFRAME_CSS)
            except Exception:
                inner_frames = []
            inner = inner_frames[0] if inner_frames else None
            if not inner:
                time.sleep(0.2)
        if not inner:
            raise RuntimeError("Inner Assist iframe with has_parent=true not found")
        if not _switch_to_frame(driver, inner):
            raise RuntimeError("Failed to switch to inner iframe")

        reload_ok = bool(driver.find_elements(By.CSS_SELECTOR, SERVICE_TYPE_RELOAD_CSS))
        input_ok = _find_service_type_input(driver, require_interactable=True) is not None
        if not reload_ok or not input_ok:
            raise RuntimeError("Capability gate failed: reload anchor or service input missing")

        # Sentinel select to confirm this context renders variant inputs.
        sentinel_ok = _select_service_type_option(
            driver,
            "(SIL) Active Night-Time",
            "(SIL) Active Night-Time",
            recorder,
            allow_contains=True,
        )
        if not sentinel_ok:
            # Last fallback: open list and accept first option.
            input_el = _find_service_type_input(driver, require_interactable=True)
            if input_el:
                _clear_and_type(input_el, "")
                _wait_for_listbox_open(driver, timeout=3.0)
                input_el.send_keys(Keys.ENTER)
                _wait_for_combobox_closed(input_el, timeout=4.0)

        day_rate_ok = bool(driver.find_elements(By.CSS_SELECTOR, "input[data-cy='day_rate-input']"))
        day_code_ok = bool(driver.find_elements(By.CSS_SELECTOR, "input[data-cy='day_code-input']"))
        if not day_rate_ok or not day_code_ok:
            raise RuntimeError("Capability gate failed: day_rate/day_code inputs absent after sentinel select")

        recorder.checker(
            "open_assist",
            "PASS",
            CHECKER_VARIANTS_EDITOR_ROUTE,
            "Switched to TP1 Add Appointment nested editor iframes",
            client_id=str(client_id),
            url=_normalize_text(getattr(driver, "current_url", "")),
        )
        recorder.checker(
            "capability_gate",
            "PASS",
            CHECKER_VARIANTS_EDITOR_CAPABILITY,
            "Variants-capable editor validated",
            client_id=str(client_id),
            url=_normalize_text(getattr(driver, "current_url", "")),
        )
        return True
    except Exception as exc:
        screenshot, html = _capture_assist_failure_artifacts(
            driver,
            recorder,
            prefix=f"switch_variants_editor_{client_id}",
            reason=str(exc),
        )
        recorder.checker(
            "open_assist",
            "FAIL",
            CHECKER_VARIANTS_EDITOR_ROUTE,
            str(exc),
            client_id=str(client_id),
            url=_normalize_text(getattr(driver, "current_url", "")),
            screenshot=screenshot,
            html=html,
        )
        return False


def _open_assist_appointments_new(
    driver,
    recorder: DiagnosticsRecorder,
    probe_client_id: str = "",
) -> bool:
    """Compatibility wrapper: open only via TP1 nested-iframe route."""
    return _switch_to_variants_capable_editor(driver, str(probe_client_id or "").strip(), recorder)


def _collect_assist_options(driver, recorder: DiagnosticsRecorder) -> List[Dict[str, str]]:
    """Collect Service Type options from listbox in inner Assist editor."""
    options: List[Dict[str, str]] = []
    input_el = _find_service_type_input(driver, require_interactable=True)
    if not input_el:
        return options

    try:
        input_el.click()
    except Exception:
        pass
    _wait_for_listbox_open(driver, timeout=4.0)

    seen = set()
    try:
        option_elems = driver.find_elements(By.XPATH, SERVICE_TYPE_OPTION_XPATH)
    except Exception:
        option_elems = []
    for elem in option_elems:
        label = _normalize_text(getattr(elem, "text", ""))
        if label and label not in seen:
            seen.add(label)
            options.append({"label": label, "value": label})

    try:
        input_el.send_keys(Keys.ESCAPE)
    except Exception:
        pass
    _wait_for_combobox_closed(input_el, timeout=3.0)

    recorder.event(
        "INFO",
        "collect_options",
        "OPTIONS_COLLECTED",
        f"Collected {len(options)} Service Type options from listbox",
        option_count=len(options),
    )
    return options


# =============================================================================
# SERVICE TYPE SELECTION WITH RETRIES
# =============================================================================


def _select_assist_option_with_retries(
    driver,
    target_value: str,
    target_label: str,
    recorder: DiagnosticsRecorder,
    max_retries: int = 3,
) -> bool:
    """Select by typing + Enter first, then exact option click fallback."""
    _ = target_value  # options carry visible text only in this editor.
    full_label = _normalize_text(target_label)
    if not full_label:
        return False
    tokens = full_label.split()
    strategies = [
        ("full_label", full_label, False),
        ("distinctive_substring", " ".join(tokens[: min(4, len(tokens))]), True),
        ("prefix", full_label[:40], True),
    ]
    for attempt in range(max_retries):
        for strategy_name, query, allow_contains in strategies:
            try:
                success = _select_service_type_option(
                    driver,
                    query,
                    full_label,
                    recorder,
                    allow_contains=allow_contains,
                )
                if success:
                    recorder.event(
                        "INFO",
                        "select_option",
                        "SERVICE_TYPE_SELECTED",
                        f"Selected '{full_label}' via {strategy_name}",
                        context={"attempt": attempt + 1, "strategy": strategy_name},
                    )
                    return True
            except Exception as exc:
                recorder.event(
                    "WARN",
                    "select_option",
                    "SELECT_STRATEGY_FAILED",
                    str(exc),
                    context={"attempt": attempt + 1, "strategy": strategy_name},
                )
                continue
    recorder.event(
        "ERROR",
        "select_option",
        "SELECT_FAILED_ALL_STRATEGIES",
        f"Could not select '{full_label}'",
    )
    return False


def _load_service_types_from_reference_index(
    recorder: DiagnosticsRecorder,
) -> Dict[str, str]:
    """Load Service Type ID -> label mapping from latest reference export."""
    service_types: Dict[str, str] = {}
    candidates = [
        line_item_paths.reference_latest_csv(),
        line_item_paths.reference_latest_xlsx(),
    ]

    for source_path in candidates:
        if not source_path.exists():
            continue
        try:
            rows: List[Dict[str, str]] = []
            if source_path.suffix.lower() == ".csv":
                with open(source_path, "r", newline="", encoding="utf-8") as fh:
                    rows = list(csv.DictReader(fh))
            else:
                wb = load_workbook(source_path, read_only=True, data_only=True)
                ws = wb.active
                header = []
                for cell in ws[1]:
                    header.append(_normalize_text(cell.value))
                for values in ws.iter_rows(min_row=2, values_only=True):
                    row = {}
                    for idx, value in enumerate(values):
                        if idx < len(header) and header[idx]:
                            row[header[idx]] = _normalize_text(value)
                    rows.append(row)

            if not rows:
                continue

            for row in rows:
                id_value = (
                    row.get("ID")
                    or row.get("Service Type ID")
                    or row.get("ServiceTypeID")
                    or row.get("ServiceType Id")
                    or ""
                )
                label_value = (
                    row.get("Service Type")
                    or row.get("Name")
                    or row.get("ServiceType")
                    or row.get("Service Type Label")
                    or ""
                )
                id_value = _normalize_text(id_value)
                label_value = _normalize_text(label_value)
                if id_value and label_value and id_value not in service_types:
                    service_types[id_value] = label_value

            if service_types:
                recorder.event(
                    "INFO",
                    "service_type_index",
                    "REFERENCE_INDEX_LOADED",
                    f"Loaded {len(service_types)} Service Types from {source_path.name}",
                    option_count=len(service_types),
                    context={"path": str(source_path)},
                )
                return service_types
        except Exception as exc:
            recorder.event(
                "WARN",
                "service_type_index",
                "REFERENCE_INDEX_LOAD_FAILED",
                f"Could not load {source_path}: {exc}",
                context={"path": str(source_path)},
            )
            continue

    recorder.event(
        "WARN",
        "service_type_index",
        "REFERENCE_INDEX_MISSING",
        "No reference Service Type export found; falling back to UI options",
    )
    return service_types


def _build_variant_record(
    *,
    service_type_id: str,
    service_type_label: str,
    variant: Dict[str, str],
    probe_client_id: str,
    source_url: str,
) -> Dict[str, str]:
    code_value = variant.get("Code", "")
    return {
        "Parent Service Type ID": service_type_id,
        "Parent Service Type Label": service_type_label,
        "Parent Service Type": service_type_label,
        "Service Type ID": service_type_id,
        "Service Variant Label": variant.get("Service Variant Label", ""),
        "Rate": variant.get("Rate", ""),
        "Rate (Raw)": variant.get("Rate (Raw)", ""),
        "Code": code_value,
        "Code (Raw)": variant.get("Code (Raw)", ""),
        "Item Number": code_value,
        "Unit": variant.get("Unit", ""),
        "Status": "PASS",
        "Error Reason": "",
        "Conflict": "UNKNOWN",
        "Conflict Detail": "",
        "Probe Client ID": probe_client_id,
        "Source URL": source_url,
        "Captured At (UTC)": _utc_now(),
    }


def _build_failure_record(
    *,
    service_type_id: str,
    service_type_label: str,
    probe_client_id: str,
    source_url: str,
    reason: str,
) -> Dict[str, str]:
    return {
        "Parent Service Type ID": service_type_id,
        "Parent Service Type Label": service_type_label,
        "Parent Service Type": service_type_label,
        "Service Type ID": service_type_id,
        "Service Variant Label": "",
        "Rate": "",
        "Rate (Raw)": "",
        "Code": "",
        "Code (Raw)": "",
        "Item Number": "",
        "Unit": "",
        "Status": "FAIL",
        "Error Reason": _normalize_text(reason),
        "Conflict": "N/A",
        "Conflict Detail": "",
        "Probe Client ID": probe_client_id,
        "Source URL": source_url,
        "Captured At (UTC)": _utc_now(),
    }


# =============================================================================
# CONFLICT DETECTION
# =============================================================================


def _detect_conflicts(variants: List[Dict[str, str]]) -> Tuple[List[Dict], List[Dict]]:
    """
    Detect conflicting variants across probe clients.
    Returns (clean_rows, conflict_rows) where conflict rows have
    "Conflict" = "YES" and conflict detail.
    """
    by_key: Dict[Tuple[str, str], List[Dict[str, str]]] = {}
    clean_rows: List[Dict[str, str]] = []
    conflict_rows: List[Dict[str, str]] = []

    for v in variants:
        if _normalize_text(v.get("Status", "PASS")).upper() == "FAIL":
            v["Conflict"] = "N/A"
            v["Conflict Detail"] = ""
            clean_rows.append(v)
            continue
        key = (v.get("Parent Service Type ID", ""), v.get("Service Variant Label", ""))
        if key not in by_key:
            by_key[key] = []
        by_key[key].append(v)

    for key, rows in by_key.items():
        rates = set(r.get("Rate", "") for r in rows if r.get("Rate"))
        codes = set(r.get("Code", "") for r in rows if r.get("Code"))

        if len(rates) > 1 or len(codes) > 1:
            for r in rows:
                r["Conflict"] = "YES"
                r["Conflict Detail"] = f"rates={rates} | codes={codes}"
            conflict_rows.extend(rows)
        else:
            for r in rows:
                r["Conflict"] = "NO"
                r["Conflict Detail"] = ""
            clean_rows.extend(rows)

    return clean_rows, conflict_rows


# =============================================================================
# MAIN EXTRACTION PIPELINE
# =============================================================================


def extract_service_type_variants(
    *,
    headless: bool = True,
    probe_client_ids: Union[str, List[str]],
    on_progress: ProgressCallback = None,
    on_event: EventCallback = None,
    on_row: RowCallback = None,
    resume: bool = True,
    force_refresh: bool = False,
    smoke_mode: bool = False,
    smoke_service_type_id: str = "",
    smoke_service_type_label: str = "",
) -> Dict[str, object]:
    """
    Extract all Service Type variants from Assist appointments form.

    Args:
        headless: Run browser headless
        probe_client_ids: Client ID(s) to probe (str or list)
        on_progress: Progress callback
        on_event: Event callback
        on_row: Row callback (called for each variant row)
        resume: Resume from checkpoint if exists
        force_refresh: Reprocess all even if checkpointed
        smoke_mode: Restrict extraction to a single Service Type
        smoke_service_type_id: Optional Service Type ID to use in smoke mode
        smoke_service_type_label: Optional Service Type label to use in smoke mode

    Returns:
        Summary dict with stats
    """
    # Normalize probe_client_ids
    if isinstance(probe_client_ids, str):
        probe_client_ids = [probe_client_ids]
    probe_client_ids = [str(cid).strip() for cid in probe_client_ids if cid]
    if not probe_client_ids:
        raise RuntimeError("At least one probe_client_id is required for TP1 Add Appointment extraction.")

    # Create run ID
    run_id = line_item_paths.make_run_id("VARIANTS")
    _emit(f"Run ID: {run_id}", on_progress)

    # Setup diagnostics
    paths = line_item_paths.get_variant_paths(run_id)
    diag_dir = paths["diagnostics_dir"]
    recorder = DiagnosticsRecorder(run_id, diag_dir, on_event=on_event)

    # Load checkpoint if resuming
    checkpoint = None
    if resume:
        checkpoint = _load_checkpoint(run_id)
        if checkpoint:
            _emit(f"Resumed from checkpoint: {len(checkpoint.get('processed_service_type_ids', []))} processed", on_progress)

    if not checkpoint:
        checkpoint = {
            "run_id": run_id,
            "started_at": _utc_now(),
            "probe_clients": probe_client_ids,
            "processed_service_type_ids": [],
            "failed_service_type_ids": [],
            "total_variant_rows": 0,
            "status": "in_progress",
        }

    processed_ids = set(checkpoint.get("processed_service_type_ids", []))
    failed_ids = set(checkpoint.get("failed_service_type_ids", []))
    all_variants: List[Dict[str, str]] = []
    all_service_types: Dict[str, str] = {}

    driver = None
    fatal_error = ""
    try:
        # Ensure Chrome driver
        ensure_credentials()
        line_item_paths.ensure_structure()
        download_dir = line_item_paths.downloads_dir()
        driver = build_chrome_driver(headless=headless, download_dir=download_dir)
        login(driver)

        # Navigate to Assist
        _emit("Opening Assist appointments form...", on_progress)
        preferred_probe = probe_client_ids[0] if probe_client_ids else ""
        if not _open_assist_appointments_new(driver, recorder, preferred_probe):
            raise RuntimeError("Failed to open Assist appointment editor with interactable Service Type combobox.")

        # Collect Service Types from exported index first, fallback to UI options.
        all_service_types = _load_service_types_from_reference_index(recorder)
        if not all_service_types:
            options = _collect_assist_options(driver, recorder)
            for opt in options:
                value = _normalize_text(opt.get("value") or "")
                label = _normalize_text(opt.get("label") or "")
                if value and label and value not in all_service_types:
                    all_service_types[value] = label

        if not all_service_types:
            raise RuntimeError("No Service Types available from reference index or Assist combobox options.")

        if smoke_mode:
            selected: Dict[str, str] = {}
            target_id = _normalize_text(smoke_service_type_id)
            target_label = _normalize_text(smoke_service_type_label).lower()
            if target_id and target_id in all_service_types:
                selected[target_id] = all_service_types[target_id]
            elif target_label:
                for st_id, st_label in sorted(all_service_types.items()):
                    if target_label in st_label.lower():
                        selected[st_id] = st_label
                        break
            if not selected:
                first_id = next(iter(sorted(all_service_types.keys())))
                selected[first_id] = all_service_types[first_id]
            all_service_types = selected
            recorder.event(
                "INFO",
                "smoke_mode",
                "SMOKE_MODE_ACTIVE",
                f"Smoke mode active for {len(all_service_types)} Service Type(s)",
                context=all_service_types,
            )

        _emit(f"Total Service Types queued: {len(all_service_types)}", on_progress)

        # Extract variants for each Service Type
        current_probe_context = preferred_probe
        for st_value, st_label in sorted(all_service_types.items(), key=lambda kv: kv[1].lower()):
            if force_refresh:
                processed_flag = False
            else:
                processed_flag = st_value in processed_ids

            if processed_flag:
                _emit(f"Skipping {st_label} (already processed)", on_progress)
                continue

            service_pass = False
            probes_for_type = probe_client_ids if len(probe_client_ids) > 1 else [preferred_probe]
            for probe_id in probes_for_type:
                source_url = _normalize_text(getattr(driver, "current_url", ""))
                try:
                    _emit(f"Extracting variants for {st_label} (probe {probe_id})...", on_progress)

                    if probe_id != current_probe_context or len(probes_for_type) > 1:
                        if not _switch_to_variants_capable_editor(driver, probe_id, recorder):
                            reason = "Unable to open variants-capable nested iframe editor for probe."
                            failure_row = _build_failure_record(
                                service_type_id=st_value,
                                service_type_label=st_label,
                                probe_client_id=probe_id,
                                source_url=_normalize_text(getattr(driver, "current_url", "")) or source_url,
                                reason=reason,
                            )
                            all_variants.append(failure_row)
                            _append_to_checkpoint_csv(run_id, [failure_row])
                            if on_row:
                                on_row(failure_row)
                            checkpoint["total_variant_rows"] = len(all_variants)
                            _update_checkpoint(run_id, checkpoint)
                            continue
                        current_probe_context = probe_id

                    before_fp = _read_variant_fingerprint(driver)
                    success = _select_assist_option_with_retries(
                        driver,
                        st_value,
                        st_label,
                        recorder,
                    )
                    if not success:
                        reason = "Service Type selection failed after fallback strategies."
                        screenshot, html = _capture_assist_failure_artifacts(
                            driver,
                            recorder,
                            prefix=f"service_type_select_fail_{st_value}_{probe_id}",
                            reason=reason,
                        )
                        recorder.checker(
                            "service_type_select",
                            "FAIL",
                            CHECKER_SERVICE_TYPE_SELECT,
                            reason,
                            client_id=probe_id,
                            url=_normalize_text(getattr(driver, "current_url", "")),
                            selector=SERVICE_TYPE_INPUT_XPATH,
                            screenshot=screenshot,
                            html=html,
                        )
                        failure_row = _build_failure_record(
                            service_type_id=st_value,
                            service_type_label=st_label,
                            probe_client_id=probe_id,
                            source_url=_normalize_text(getattr(driver, "current_url", "")) or source_url,
                            reason=reason,
                        )
                        all_variants.append(failure_row)
                        _append_to_checkpoint_csv(run_id, [failure_row])
                        if on_row:
                            on_row(failure_row)
                        checkpoint["total_variant_rows"] = len(all_variants)
                        _update_checkpoint(run_id, checkpoint)
                        continue

                    if not _wait_for_fingerprint_change(driver, before_fp, timeout=6.0):
                        reason = "Variant fingerprint did not change after Service Type selection."
                        screenshot, html = _capture_assist_failure_artifacts(
                            driver,
                            recorder,
                            prefix=f"stale_fingerprint_{st_value}_{probe_id}",
                            reason=reason,
                        )
                        recorder.checker(
                            "extract_variants",
                            "FAIL",
                            CHECKER_VARIANT_TABLE_EMPTY,
                            reason,
                            client_id=probe_id,
                            url=_normalize_text(getattr(driver, "current_url", "")),
                            screenshot=screenshot,
                            html=html,
                        )
                        failure_row = _build_failure_record(
                            service_type_id=st_value,
                            service_type_label=st_label,
                            probe_client_id=probe_id,
                            source_url=_normalize_text(getattr(driver, "current_url", "")) or source_url,
                            reason=reason,
                        )
                        all_variants.append(failure_row)
                        _append_to_checkpoint_csv(run_id, [failure_row])
                        if on_row:
                            on_row(failure_row)
                        checkpoint["total_variant_rows"] = len(all_variants)
                        _update_checkpoint(run_id, checkpoint)
                        continue

                    variants = _extract_variant_table_rows(driver, recorder)
                    if len(variants) != 6:
                        reason = f"Expected 6 variant pairs, got {len(variants)}."
                        screenshot, html = _capture_assist_failure_artifacts(
                            driver,
                            recorder,
                            prefix=f"variant_pairs_missing_{st_value}_{probe_id}",
                            reason=reason,
                        )
                        recorder.checker(
                            "extract_variants",
                            "FAIL",
                            CHECKER_VARIANT_TABLE_MISSING,
                            reason,
                            client_id=probe_id,
                            url=_normalize_text(getattr(driver, "current_url", "")),
                            screenshot=screenshot,
                            html=html,
                        )
                        failure_row = _build_failure_record(
                            service_type_id=st_value,
                            service_type_label=st_label,
                            probe_client_id=probe_id,
                            source_url=_normalize_text(getattr(driver, "current_url", "")) or source_url,
                            reason=reason,
                        )
                        all_variants.append(failure_row)
                        _append_to_checkpoint_csv(run_id, [failure_row])
                        if on_row:
                            on_row(failure_row)
                        checkpoint["total_variant_rows"] = len(all_variants)
                        _update_checkpoint(run_id, checkpoint)
                        continue

                    variant_records: List[Dict[str, str]] = []
                    for variant in variants:
                        record = _build_variant_record(
                            service_type_id=st_value,
                            service_type_label=st_label,
                            variant=variant,
                            probe_client_id=probe_id,
                            source_url=_normalize_text(getattr(driver, "current_url", "")),
                        )
                        variant_records.append(record)
                        if on_row:
                            on_row(record)

                    _append_to_checkpoint_csv(run_id, variant_records)
                    all_variants.extend(variant_records)
                    checkpoint["total_variant_rows"] = len(all_variants)
                    _update_checkpoint(run_id, checkpoint)
                    service_pass = True
                    _emit(f"  → extracted {len(variants)} variants (probe {probe_id})", on_progress)
                except Exception as e:
                    reason = f"Exception during Service Type extraction: {e}"
                    screenshot, html = _capture_assist_failure_artifacts(
                        driver,
                        recorder,
                        prefix=f"service_type_exception_{st_value}_{probe_id}",
                        reason=reason,
                    )
                    recorder.event(
                        "ERROR",
                        "extract_service_type",
                        "EXCEPTION",
                        reason,
                        client_id=probe_id,
                        context={"service_type": st_label},
                    )
                    recorder.checker(
                        "extract_service_type",
                        "FAIL",
                        "CHK_SERVICE_TYPE_EXCEPTION",
                        reason,
                        client_id=probe_id,
                        url=_normalize_text(getattr(driver, "current_url", "")),
                        screenshot=screenshot,
                        html=html,
                    )
                    failure_row = _build_failure_record(
                        service_type_id=st_value,
                        service_type_label=st_label,
                        probe_client_id=probe_id,
                        source_url=_normalize_text(getattr(driver, "current_url", "")) or source_url,
                        reason=reason,
                    )
                    all_variants.append(failure_row)
                    _append_to_checkpoint_csv(run_id, [failure_row])
                    if on_row:
                        on_row(failure_row)
                    checkpoint["total_variant_rows"] = len(all_variants)
                    _update_checkpoint(run_id, checkpoint)
                    continue

            if service_pass:
                processed_ids.add(st_value)
                checkpoint["processed_service_type_ids"] = list(processed_ids)
                if st_value in failed_ids:
                    failed_ids.remove(st_value)
            else:
                failed_ids.add(st_value)
            checkpoint["failed_service_type_ids"] = list(failed_ids)
            checkpoint["total_variant_rows"] = len(all_variants)
            _update_checkpoint(run_id, checkpoint)

        # Detect conflicts
        _emit("Detecting conflicts...", on_progress)
        clean_variants, conflict_variants = _detect_conflicts(all_variants)

        # Write outputs
        _emit("Writing outputs...", on_progress)

        # Latest CSV/XLSX
        _write_csv(all_variants, VARIANT_COLUMNS, paths["latest_csv"])
        _write_xlsx(all_variants, VARIANT_COLUMNS, paths["latest_xlsx"])
        _apply_xlsx_merged_cells(paths["latest_xlsx"])

        # Snapshot CSV/XLSX
        _write_csv(all_variants, VARIANT_COLUMNS, paths["snapshot_csv"])
        _write_xlsx(all_variants, VARIANT_COLUMNS, paths["snapshot_xlsx"])
        _apply_xlsx_merged_cells(paths["snapshot_xlsx"])

        # Conflicts CSV
        if conflict_variants:
            conflict_cols = VARIANT_COLUMNS
            _write_csv(conflict_variants, conflict_cols, paths["conflicts_csv"])

        # Finalize checkpoint
        checkpoint["status"] = "completed"
        checkpoint["completed_at"] = _utc_now()
        checkpoint["processed_service_type_ids"] = list(processed_ids)
        checkpoint["failed_service_type_ids"] = list(failed_ids)
        checkpoint["total_variant_rows"] = len(all_variants)
        _update_checkpoint(run_id, checkpoint)

        summary = {
            "run_id": run_id,
            "total_service_types": len(all_service_types),
            "processed_service_types": len(processed_ids),
            "failed_service_types": len(failed_ids),
            "total_variant_rows": len(all_variants),
            "clean_variants": len(clean_variants),
            "conflict_variants": len(conflict_variants),
            "smoke_mode": smoke_mode,
            "output_paths": {
                "latest_csv": str(paths["latest_csv"]),
                "latest_xlsx": str(paths["latest_xlsx"]),
                "snapshot_csv": str(paths["snapshot_csv"]),
                "snapshot_xlsx": str(paths["snapshot_xlsx"]),
                "conflicts_csv": str(paths["conflicts_csv"]),
                "diagnostics_dir": str(diag_dir),
            },
        }

        _emit(f"Extraction complete! Summary: {summary}", on_progress)

        return summary
    except Exception as exc:
        fatal_error = str(exc)
        recorder.event(
            "ERROR",
            "extract_run",
            "RUN_FAILED",
            fatal_error,
            context={"traceback": traceback.format_exc()},
        )
        if driver:
            screenshot, html = _capture_assist_failure_artifacts(
                driver,
                recorder,
                prefix="run_failure",
                reason=fatal_error,
            )
            recorder.checker(
                "extract_run",
                "FAIL",
                "CHK_VARIANT_EXTRACTION_RUN",
                fatal_error,
                url=_normalize_text(getattr(driver, "current_url", "")),
                screenshot=screenshot,
                html=html,
            )
        checkpoint["status"] = "failed"
        checkpoint["completed_at"] = _utc_now()
        checkpoint["error"] = fatal_error
        checkpoint["processed_service_type_ids"] = list(processed_ids)
        checkpoint["failed_service_type_ids"] = list(failed_ids)
        checkpoint["total_variant_rows"] = len(all_variants)
        _update_checkpoint(run_id, checkpoint)
        raise RuntimeError(f"Service Type variant extraction failed: {fatal_error}") from exc
    finally:
        try:
            recorder.save()
        except Exception as save_exc:
            log_message(f"Warning: could not persist diagnostics recorder output: {save_exc}")
        if driver:
            try:
                driver.quit()
            except Exception:
                pass


# =============================================================================
# LEGACY STUBS (backward compatibility for UI imports)
# =============================================================================


def load_discovery_latest(path: Path = None) -> List[Dict[str, str]]:
    """Load latest variant data. Legacy name kept for backward compatibility."""
    if path is None:
        paths = line_item_paths.get_variant_paths(
            line_item_paths.make_run_id("LEGACY")
        )
        path = paths["latest_csv"]

    if not path.exists():
        return []

    try:
        rows = []
        with open(path, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        return rows
    except Exception:
        return []


def discover_appointment_item_numbers(**kwargs) -> Dict[str, object]:
    """Legacy stub — redirects to extract_service_type_variants().

    Translates old parameter names to new ones so existing UI code
    continues to work without modification.
    """
    # Map old parameter names to new ones
    new_kwargs = {}
    new_kwargs["headless"] = kwargs.get("headless", True)

    probe = kwargs.get("probe_client_id") or kwargs.get("probe_client_ids")
    if probe:
        new_kwargs["probe_client_ids"] = probe
    else:
        new_kwargs["probe_client_ids"] = []

    new_kwargs["on_progress"] = kwargs.get("on_progress")
    new_kwargs["on_event"] = kwargs.get("on_event")
    new_kwargs["on_row"] = kwargs.get("on_row")
    new_kwargs["resume"] = kwargs.get("resume", True)
    new_kwargs["force_refresh"] = kwargs.get("force_refresh", False)
    new_kwargs["smoke_mode"] = kwargs.get("smoke_mode", False)
    new_kwargs["smoke_service_type_id"] = kwargs.get("smoke_service_type_id", "")
    new_kwargs["smoke_service_type_label"] = kwargs.get("smoke_service_type_label", "")

    result = extract_service_type_variants(**new_kwargs)

    # Map new result keys to old ones for backward compatibility
    output_paths = result.get("output_paths", {})
    result["row_count"] = result.get("total_variant_rows", 0)
    result["rows"] = []
    result["diagnostics_folder"] = output_paths.get("diagnostics_dir", "")
    result["discovery_latest_csv"] = output_paths.get("latest_csv", "")
    result["discovery_latest_xlsx"] = output_paths.get("latest_xlsx", "")
    result["output_root"] = str(Path(output_paths.get("latest_csv", "")).parent) if output_paths.get("latest_csv") else ""

    return result


def run_service_type_merge(discovered_rows: List[Dict[str, str]], **kwargs) -> Dict[str, object]:
    """Legacy stub — merge is no longer needed.

    The new variant extractor already produces complete Rate+Code data
    directly from the Assist table. This stub returns an empty result
    to keep the UI merge button from crashing.
    """
    progress = kwargs.get("progress")
    if progress:
        progress("Merge is no longer needed — variant extraction already includes Rate+Code per variant.")

    return {
        "enriched_count": 0,
        "unmatched_count": 0,
        "enriched_rows": [],
        "enriched_latest_csv": "",
        "unmatched_latest_csv": "",
    }
